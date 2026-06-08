import aiosqlite
import config

_DDL = """
CREATE TABLE IF NOT EXISTS leads (
    id         INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id    INTEGER NOT NULL,
    username   TEXT,
    full_name  TEXT,
    intent     TEXT,
    product    TEXT,
    leads      TEXT,
    budget     TEXT,
    timeline   TEXT,
    bot_reaction TEXT,
    demo_pick  TEXT,
    utm_source TEXT,
    utm_medium TEXT,
    completed  INTEGER DEFAULT 0,
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS users (
    user_id    INTEGER PRIMARY KEY,
    username   TEXT,
    full_name  TEXT,
    first_seen DATETIME DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS followups (
    id      INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    step    INTEGER NOT NULL DEFAULT 0,
    send_at DATETIME NOT NULL,
    sent    INTEGER DEFAULT 0
);
"""

FOLLOWUP_STEPS = [1, 24, 72]


async def init_db():
    import os
    os.makedirs(os.path.dirname(config.DB_PATH), exist_ok=True)
    async with aiosqlite.connect(config.DB_PATH) as db:
        for stmt in _DDL.strip().split(";"):
            s = stmt.strip()
            if s:
                await db.execute(s)
        await db.commit()


async def upsert_user(user_id, username, full_name):
    async with aiosqlite.connect(config.DB_PATH) as db:
        await db.execute(
            "INSERT OR IGNORE INTO users (user_id,username,full_name) VALUES(?,?,?)",
            (user_id, username, full_name))
        await db.execute(
            "UPDATE users SET username=?,full_name=? WHERE user_id=?",
            (username, full_name, user_id))
        await db.commit()


async def save_lead(user_id, username, full_name, answers, utm, completed=False):
    cols = ["user_id","username","full_name","utm_source","utm_medium","completed"]
    vals = [user_id, username, full_name,
            utm.get("utm_source"), utm.get("utm_medium"), int(completed)]
    allowed = {"intent","product","leads","budget","timeline","bot_reaction","demo_pick"}
    for k, v in answers.items():
        if k in allowed:
            cols.append(k)
            vals.append(v)
    ph = ",".join("?" * len(cols))
    async with aiosqlite.connect(config.DB_PATH) as db:
        cur = await db.execute(
            f"INSERT INTO leads ({','.join(cols)}) VALUES ({ph})", vals)
        await db.commit()
        return cur.lastrowid


async def schedule_followups(user_id):
    from datetime import datetime, timedelta
    async with aiosqlite.connect(config.DB_PATH) as db:
        await db.execute("DELETE FROM followups WHERE user_id=?", (user_id,))
        for step, hours in enumerate(FOLLOWUP_STEPS):
            send_at = (datetime.utcnow() + timedelta(hours=hours)).isoformat()
            await db.execute(
                "INSERT INTO followups (user_id,step,send_at) VALUES(?,?,?)",
                (user_id, step, send_at))
        await db.commit()


async def cancel_followup(user_id):
    async with aiosqlite.connect(config.DB_PATH) as db:
        await db.execute("DELETE FROM followups WHERE user_id=?", (user_id,))
        await db.commit()


async def get_pending_followups():
    from datetime import datetime
    now = datetime.utcnow().isoformat()
    async with aiosqlite.connect(config.DB_PATH) as db:
        cur = await db.execute(
            "SELECT id,user_id,step FROM followups WHERE sent=0 AND send_at<=? ORDER BY step",
            (now,))
        return await cur.fetchall()


async def mark_followup_sent(fid):
    async with aiosqlite.connect(config.DB_PATH) as db:
        await db.execute("UPDATE followups SET sent=1 WHERE id=?", (fid,))
        await db.commit()


async def get_all_user_ids():
    async with aiosqlite.connect(config.DB_PATH) as db:
        cur = await db.execute("SELECT user_id FROM users")
        return [r[0] for r in await cur.fetchall()]


async def export_leads_xlsx(path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    async with aiosqlite.connect(config.DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        cur = await db.execute("SELECT * FROM leads ORDER BY id DESC")
        rows = await cur.fetchall()
    if not rows:
        return 0
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лиды"
    headers = list(rows[0].keys())
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="635BFF")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(15, len(h)+4)
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, val)
    wb.save(path)
    return len(rows)


async def get_stats():
    async with aiosqlite.connect(config.DB_PATH) as db:
        async def one(sql, *a):
            cur = await db.execute(sql, a)
            return (await cur.fetchone())[0]
        total     = await one("SELECT COUNT(*) FROM leads")
        completed = await one("SELECT COUNT(*) FROM leads WHERE completed=1")
        today     = await one("SELECT COUNT(*) FROM leads WHERE date(created_at)=date('now')")
        users     = await one("SELECT COUNT(*) FROM users")
        cur = await db.execute(
            "SELECT product,COUNT(*) FROM leads WHERE product IS NOT NULL "
            "GROUP BY product ORDER BY 2 DESC")
        needs = await cur.fetchall()
    return {
        "total": total, "completed": completed, "today": today, "users": users,
        "conversion": round(completed/total*100,1) if total else 0,
        "needs": needs,
    }
